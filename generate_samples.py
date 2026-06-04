"""
Generate three sample IC Resolve upload files and save to DUMMY DATA folder.
Run from the ic-resolve directory: python generate_samples.py
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUT_DIR = r"C:\Users\kanne\Documents\DUMMY DATA"
os.makedirs(OUT_DIR, exist_ok=True)

PERIOD_LABEL = "March 2026"
PERIOD_CODE  = "2026-03"

# ── helpers ──────────────────────────────────────────────────────────────────

DARK   = "1e1b4b"
MID    = "4f46e5"
LIGHT  = "e0e7ff"
HEADER = "f8f9ff"

def _hdr(ws, row, col, value, bold=False, bg=None, align="left"):
    c = ws.cell(row, col, value)
    c.font = Font(bold=bold, color="FFFFFF" if bg and bg not in (LIGHT, HEADER) else "1e1b4b")
    if bg:
        c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    return c

def _val(ws, row, col, value, num_fmt=None, bold=False, align="left"):
    c = ws.cell(row, col, value)
    if num_fmt:
        c.number_format = num_fmt
    if bold:
        c.font = Font(bold=True)
    c.alignment = Alignment(horizontal=align)
    return c

def _set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _thin_border():
    s = Side(style="thin", color="d1d5db")
    return Border(left=s, right=s, top=s, bottom=s)

def _apply_borders(ws, min_row, max_row, min_col, max_col):
    b = _thin_border()
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = b


# ── 1. TRIAL BALANCE ─────────────────────────────────────────────────────────
#
# Pairs and expected outcomes:
#   NXR-UK ↔ NXR-US   Management Fee      → MATCHED          (diff < $1k)
#   NXR-UK ↔ NXR-DE   IT Services         → FX DIFFERENCE    (NXR-DE used stale rate)
#   NXR-UK ↔ NXR-SG   Intercompany Loan   → MISSING POSTING  (NXR-SG silent)
#   NXR-DE ↔ NXR-AU   Distribution Fee    → AMOUNT DIFF      ($35k gap)
#   NXR-US ↔ NXR-SG   Sales Commission    → MATCHED          (diff < $1k)

def make_trial_balance():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 36

    # Row 1 — title
    ws.merge_cells("A1:Q1")
    _hdr(ws, 1, 1, "Nexora International Group  |  TRIAL BALANCE EXTRACT", bold=True, bg=DARK)

    # Row 2 — metadata
    ws.merge_cells("A2:Q2")
    _hdr(ws, 2, 1,
         f"Period: {PERIOD_LABEL}  |  Scenario: Actual  |  Version: Preliminary  |  Reporting Currency: USD",
         bg=MID)

    # Row 3 — section labels
    for col, (label, span_end) in enumerate([
        (1,  ("ENTITY DIMENSIONS",    5)),
        (6,  ("ACCOUNT DIMENSIONS",   9)),
        (10, ("ANALYTICAL DIMENSIONS",13)),
        (14, ("INTERCOMPANY",         15)),
        (16, ("FINANCIALS",           17)),
    ], 0):
        start, end = [(1,5),(6,9),(10,13),(14,15),(16,17)][col]
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        _hdr(ws, 3, start, [(1,5),(6,9),(10,13),(14,15),(16,17)][col] and
             ["ENTITY DIMENSIONS","ACCOUNT DIMENSIONS","ANALYTICAL DIMENSIONS","INTERCOMPANY","FINANCIALS"][col],
             bold=True, bg=LIGHT)

    # Row 4 — column headers
    cols = ["Entity Code","Entity Name","Country","Region","Consol Level",
            "Account Code","Account Name","Account Type","Flow Type",
            "Cost Centre","Cost Centre Desc","Product / BU","Project Code",
            "IC Flag","IC Counterparty","LC Amount","USD Amount"]
    for i, h in enumerate(cols, 1):
        _hdr(ws, 4, i, h, bold=True, bg=HEADER)

    # Data — IC rows only (non-IC rows omitted for brevity in this sample)
    # Format: (entity_code, entity_name, country, region, consol, acct_code, acct_name,
    #          acct_type, flow, cc, cc_desc, product, project, ic_flag, ic_cparty, lc, usd)
    rows = [
        # ── NXR-UK ──────────────────────────────────────────────────────────
        ("NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1312","IC Receivable – NXR-US","Asset","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-US",500_000,632_500),
        ("NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1311","IC Receivable – NXR-DE","Asset","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-DE",250_000,316_250),
        ("NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1600","IC Loan Receivable – NXR-SG","Asset","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-SG",1_000_000,1_265_000),

        # ── NXR-US ──────────────────────────────────────────────────────────
        # Matches NXR-UK management fee (USD amount within tolerance)
        ("NXR-US","Nexora Americas Inc","United States","Americas","L2",
         "2311","IC Payable – NXR-UK","Liability","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-UK",-632_000,-632_000),
        # Matches NXR-SG commission
        ("NXR-US","Nexora Americas Inc","United States","Americas","L2",
         "1504","IC Receivable – NXR-SG","Asset","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-SG",85_000,85_000),

        # ── NXR-DE ──────────────────────────────────────────────────────────
        # FX error: NXR-DE used rate 1.065 instead of official 1.082
        # EUR 292,250 × 1.065 = USD 311,246  (official: × 1.082 = USD 316,214)
        ("NXR-DE","Nexora Deutschland GmbH","Germany","EMEA","L2",
         "2501","IC Payable – NXR-UK","Liability","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-UK",-292_250,-311_246),
        # Amount difference vs NXR-AU
        ("NXR-DE","Nexora Deutschland GmbH","Germany","EMEA","L2",
         "1505","IC Receivable – NXR-AU","Asset","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-AU",150_000,162_300),

        # ── NXR-SG ──────────────────────────────────────────────────────────
        # Missing: no IC loan payable to NXR-UK
        # Commission payable to NXR-US — matches within tolerance
        ("NXR-SG","Nexora Asia Pacific Pte","Singapore","APAC","L2",
         "2504","IC Payable – NXR-US","Liability","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-US",-114_500,-84_979),

        # ── NXR-AU ──────────────────────────────────────────────────────────
        # Genuine amount difference vs NXR-DE (AUD 200k ≠ EUR 150k equivalent)
        ("NXR-AU","Nexora Australia Pty","Australia","APAC","L2",
         "2505","IC Payable – NXR-DE","Liability","Closing Balance",
         "CC-FIN","Finance & Treasury","Corporate",None,"Y","NXR-DE",-200_000,-127_600),
    ]

    for r, row in enumerate(rows, 5):
        for c, val in enumerate(row, 1):
            cell = _val(ws, r, c, val,
                        num_fmt="#,##0" if c in (16, 17) else None,
                        align="right" if c in (16, 17) else "left")
            if row[13] == "Y":  # IC flag
                cell.fill = PatternFill("solid", fgColor="f0fdf4" if row[7] == "Asset" else "fff7ed")

    _apply_borders(ws, 4, 4 + len(rows), 1, 17)
    _set_widths(ws, [10,22,16,8,7,10,28,10,14,8,18,10,9,6,10,14,14])

    path = os.path.join(OUT_DIR, "IC_Resolve_Trial_Balance_Sample.xlsx")
    wb.save(path)
    print(f"✓ Trial Balance   → {path}")


# ── 2. IC TRANSACTIONS ───────────────────────────────────────────────────────

def make_ic_transactions():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IC Transactions"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws.row_dimensions[4].height = 36

    ws.merge_cells("A1:S1")
    _hdr(ws, 1, 1, "Nexora International Group  |  INTERCOMPANY TRANSACTION DETAIL", bold=True, bg=DARK)
    ws.merge_cells("A2:S2")
    _hdr(ws, 2, 1,
         f"Period: {PERIOD_LABEL}  |  Scenario: Actual  |  Source: ERP Subledger Extract  |  Amounts in transaction currency",
         bg=MID)

    # Row 3 section headers (simplified)
    for start, end, label in [(1,2,"TRANSACTION"),(3,7,"ENTITY DIMENSIONS"),
                               (8,11,"ACCOUNT DIMENSIONS"),(12,15,"ANALYTICAL DIMENSIONS"),
                               (16,17,"INTERCOMPANY"),(18,19,"FINANCIALS")]:
        ws.merge_cells(start_row=3, start_column=start, end_row=3, end_column=end)
        _hdr(ws, 3, start, label, bold=True, bg=LIGHT)

    cols = ["Txn ID","Posting Date","Entity Code","Entity Name","Country","Region","Consol Level",
            "Account Code","Account Name","Account Type","Flow Type",
            "Cost Centre","Cost Ctr Desc","Product / BU","Project Code",
            "IC Flag","IC Counterparty","LC Amount","USD Amount"]
    for i, h in enumerate(cols, 1):
        _hdr(ws, 4, i, h, bold=True, bg=HEADER)

    # (txn_id, date, entity, name, country, region, consol,
    #  acct, acct_name, acct_type, flow, cc, cc_desc, product, project,
    #  ic_flag, ic_cparty, lc, usd)
    from datetime import date
    rows = [
        ("IC-2603-001", date(2026,3,5),  "NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1312","IC Receivable – NXR-US","Asset","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-US",500_000,632_500),

        ("IC-2603-002", date(2026,3,5),  "NXR-US","Nexora Americas Inc","United States","Americas","L2",
         "2311","IC Payable – NXR-UK","Liability","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-UK",-632_000,-632_000),

        ("IC-2603-003", date(2026,3,12), "NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1311","IC Receivable – NXR-DE","Asset","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-DE",250_000,316_250),

        # NXR-DE used stale rate 1.065 instead of official 1.082
        ("IC-2603-004", date(2026,3,10), "NXR-DE","Nexora Deutschland GmbH","Germany","EMEA","L2",
         "2501","IC Payable – NXR-UK","Liability","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-UK",-292_250,-311_246),

        ("IC-2603-005", date(2026,3,1),  "NXR-UK","Nexora UK Ltd","United Kingdom","EMEA","L2",
         "1600","IC Loan Receivable – NXR-SG","Asset","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-SG",1_000_000,1_265_000),
        # NXR-SG has NO matching loan payable entry (missing posting)

        ("IC-2603-006", date(2026,3,18), "NXR-DE","Nexora Deutschland GmbH","Germany","EMEA","L2",
         "1505","IC Receivable – NXR-AU","Asset","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-AU",150_000,162_300),

        ("IC-2603-007", date(2026,3,20), "NXR-AU","Nexora Australia Pty","Australia","APAC","L2",
         "2505","IC Payable – NXR-DE","Liability","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-DE",-200_000,-127_600),   # genuine amount difference — AUD 200k ≠ EUR 150k

        ("IC-2603-008", date(2026,3,25), "NXR-US","Nexora Americas Inc","United States","Americas","L2",
         "1504","IC Receivable – NXR-SG","Asset","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-SG",85_000,85_000),

        ("IC-2603-009", date(2026,3,25), "NXR-SG","Nexora Asia Pacific Pte","Singapore","APAC","L2",
         "2504","IC Payable – NXR-US","Liability","CB","CC-FIN","Finance & Treasury","Corporate",None,
         "Y","NXR-US",-114_500,-84_979),   # SGD 114,500 × 0.742 = USD 84,979 ≈ USD 85,000 ✓
    ]

    for r, row in enumerate(rows, 5):
        for c, val in enumerate(row, 1):
            _val(ws, r, c, val,
                 num_fmt="#,##0" if c in (18, 19) else
                         "YYYY-MM-DD" if c == 2 else None,
                 align="right" if c in (18, 19) else "left")

    _apply_borders(ws, 4, 4 + len(rows), 1, 19)
    _set_widths(ws, [12,12,10,22,16,8,7,10,26,10,6,8,18,10,9,6,10,13,13])

    path = os.path.join(OUT_DIR, "IC_Resolve_IC_Transactions_Sample.xlsx")
    wb.save(path)
    print(f"✓ IC Transactions → {path}")


# ── 3. FX RATES ──────────────────────────────────────────────────────────────

def make_fx_rates():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "FX Rates"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18

    ws.merge_cells("A1:F1")
    _hdr(ws, 1, 1, f"Nexora International Group  |  FX RATE TABLE  |  Period: {PERIOD_LABEL}", bold=True, bg=DARK)
    ws.merge_cells("A2:F2")
    _hdr(ws, 2, 1, "Source: Treasury / Bloomberg  |  Rate type: Period-end closing spot  |  Base: USD", bg=MID)

    cols = ["From Currency","To Currency","Rate","Effective Date","Period","Notes"]
    for i, h in enumerate(cols, 1):
        _hdr(ws, 3, i, h, bold=True, bg=HEADER)

    from datetime import date
    # Rates as X → USD (consistent with reporting currency in the TB/TX files)
    # Official EUR/USD = 1.082 — NXR-DE used 1.065, so the system should flag that
    rates = [
        ("GBP", "USD", 1.2650, date(2026,3,31), PERIOD_CODE, "Bank of England reference rate"),
        ("EUR", "USD", 1.0820, date(2026,3,31), PERIOD_CODE, "ECB reference rate — OFFICIAL"),
        ("SGD", "USD", 0.7420, date(2026,3,31), PERIOD_CODE, "MAS reference rate"),
        ("AUD", "USD", 0.6380, date(2026,3,31), PERIOD_CODE, "RBA reference rate"),
        ("JPY", "USD", 0.0067, date(2026,3,31), PERIOD_CODE, "BOJ reference rate"),
        ("CAD", "USD", 0.7310, date(2026,3,31), PERIOD_CODE, "Bank of Canada reference rate"),
        ("CHF", "USD", 1.1230, date(2026,3,31), PERIOD_CODE, "SNB reference rate"),
    ]

    for r, (frm, to, rate, eff, period, note) in enumerate(rates, 4):
        ws.cell(r, 1, frm)
        ws.cell(r, 2, to)
        c = ws.cell(r, 3, rate)
        c.number_format = "0.0000"
        c.alignment = Alignment(horizontal="right")
        d = ws.cell(r, 4, eff)
        d.number_format = "YYYY-MM-DD"
        ws.cell(r, 5, period)
        ws.cell(r, 6, note)
        if frm == "EUR":  # highlight the rate NXR-DE got wrong
            for col in range(1, 7):
                ws.cell(r, col).fill = PatternFill("solid", fgColor="fef3c7")

    _apply_borders(ws, 3, 3 + len(rates), 1, 6)
    _set_widths(ws, [14, 14, 10, 14, 10, 38])

    # Add a note below the table
    note_row = 4 + len(rates) + 1
    ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
    c = ws.cell(note_row, 1,
        "⚠  EUR/USD highlighted — NXR-DE posted at 1.065 (stale rate). "
        "IC Resolve will detect this as an FX rate error on the NXR-UK ↔ NXR-DE pair.")
    c.font = Font(italic=True, color="92400e")
    c.fill = PatternFill("solid", fgColor="fef3c7")

    path = os.path.join(OUT_DIR, "IC_Resolve_FX_Rates_Sample.xlsx")
    wb.save(path)
    print(f"✓ FX Rates        → {path}")


if __name__ == "__main__":
    make_trial_balance()
    make_ic_transactions()
    make_fx_rates()
    print(f"\nAll files saved to: {OUT_DIR}")
    print("\nExpected reconciliation results:")
    print("  NXR-UK ↔ NXR-US   Management Fee      → MATCHED          (USD 500 diff)")
    print("  NXR-UK ↔ NXR-DE   IT Services         → FX DIFFERENCE    (NXR-DE used 1.065 vs official 1.082)")
    print("  NXR-UK ↔ NXR-SG   IC Loan             → MISSING POSTING  (NXR-SG has no entry)")
    print("  NXR-DE ↔ NXR-AU   Distribution Fee    → AMOUNT DIFF      (USD 34,700 gap)")
    print("  NXR-US ↔ NXR-SG   Commission          → MATCHED          (USD 21 diff)")
