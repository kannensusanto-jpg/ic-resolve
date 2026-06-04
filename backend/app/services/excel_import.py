"""
Parse Nexora-format Excel uploads into IC Resolve data models.

Expected formats
  Trial Balance : 4-row header, data from row 5
                  Cols: Entity Code, Entity Name, Country, Region, Consol Level,
                        Account Code, Account Name, Account Type, Flow Type,
                        Cost Centre, Cost Ctr Desc, Product/BU, Project Code,
                        IC Flag, IC Counterparty, LC Amount, USD Amount
  IC Transactions: 4-row header, data from row 5
                  Cols: Txn ID, Posting Date, Entity Code, Entity Name, Country,
                        Region, Consol Level, Account Code, Account Name, Account Type,
                        Flow Type, Cost Centre, Cost Ctr Desc, Product/BU, Project Code,
                        IC Flag, IC Counterparty, LC Amount, USD Amount
  FX Rates (flexible): any layout with a header row containing recognisable column names.
                  Required: a currency column + a rate column.
                  Optional: a date column, a to-currency column (defaults to GBP), a period column.
                  Supported column name aliases — see HEADER_ALIASES below.
"""
import re
import calendar
import openpyxl
from datetime import date, datetime
from io import BytesIO
from sqlalchemy.orm import Session
from app.models import Entity, JournalEntry, TrialBalance, CloseCalendar, FXRate

CURRENCY_MAP = {
    "United Kingdom":  "GBP",
    "Germany":         "EUR",
    "United States":   "USD",
    "Singapore":       "SGD",
    "Australia":       "AUD",
    "France":          "EUR",
    "Netherlands":     "EUR",
    "Japan":           "JPY",
    "Canada":          "CAD",
}

# For reporting: treat the USD amount column as the GBP-equivalent
# (file reporting currency is USD; flag in entity name shows no GBP rate supplied)
REPORTING_CURRENCY = "USD"


def _parse_period(ws) -> str:
    """Extract period string like '2026-04' from the metadata row."""
    meta = str(ws.cell(2, 1).value or "")
    m = re.search(r"Period:\s*(\w+)\s+(\d{4})", meta)
    if m:
        months = {"January":"01","February":"02","March":"03","April":"04",
                  "May":"05","June":"06","July":"07","August":"08",
                  "September":"09","October":"10","November":"11","December":"12"}
        return f"{m.group(2)}-{months.get(m.group(1), '01')}"
    return "2026-04"


def _infer_journal_type(account_type: str, lc_amount: float) -> str:
    if account_type == "Asset":
        return "IC_RECEIVABLE"
    if account_type == "Liability":
        return "IC_PAYABLE"
    # P&L: negative = entity earns (receivable), positive = entity pays (payable)
    return "IC_RECEIVABLE" if (lc_amount or 0) < 0 else "IC_PAYABLE"


def _upsert_entity(db: Session, code: str, name: str, country: str, region: str):
    existing = db.get(Entity, code)
    currency = CURRENCY_MAP.get(country.strip(), "USD")
    if not existing:
        db.add(Entity(
            id=code,
            name=name,
            aliases=[],
            functional_currency=currency,
            region=region.strip() if region else None,
            ic_agreement_flag=True,
        ))
    return code


def import_trial_balance(db: Session, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    period = _parse_period(ws)

    # Delete existing TB rows for this period
    db.query(TrialBalance).filter(TrialBalance.period == period).delete()

    rows = list(ws.iter_rows(min_row=5, values_only=True))
    imported = skipped = 0

    for row in rows:
        if not row[0]:
            continue
        (entity_code, entity_name, country, region, consol,
         acct_code, acct_name, acct_type, flow_type,
         cc, cc_desc, product, project,
         ic_flag, ic_counterparty, lc_amount, usd_amount) = row[:17]

        if ic_flag != "Y" or not ic_counterparty:
            skipped += 1
            continue

        _upsert_entity(db, entity_code, entity_name or entity_code, country or "", region or "")

        db.add(TrialBalance(
            entity_id=entity_code,
            account_code=str(acct_code),
            account_name=str(acct_name).replace("�", "–") if acct_name else None,
            counterparty_entity_id=ic_counterparty,
            balance_local=float(lc_amount or 0),
            currency=CURRENCY_MAP.get((country or "").strip(), "USD"),
            balance_gbp=float(usd_amount or 0),  # USD treated as reporting currency
            period=period,
        ))
        imported += 1

    db.commit()
    return {"period": period, "rows_imported": imported, "rows_skipped": skipped}


def import_ic_transactions(db: Session, file_bytes: bytes) -> dict:
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    ws = wb.active
    period = _parse_period(ws)

    # Delete existing journal entries for this period
    db.query(JournalEntry).filter(JournalEntry.period == period).delete()

    rows = list(ws.iter_rows(min_row=5, values_only=True))
    imported = skipped = 0

    for row in rows:
        if not row[0] or str(row[0]).startswith("TOTAL"):
            continue
        (txn_id, posting_date, entity_code, entity_name, country, region, consol,
         acct_code, acct_name, acct_type, flow_type,
         cc, cc_desc, product, project,
         ic_flag, ic_counterparty, lc_amount, usd_amount) = row[:19]

        if ic_flag != "Y" or not ic_counterparty:
            skipped += 1
            continue

        _upsert_entity(db, entity_code, entity_name or entity_code, country or "", region or "")

        # Parse posting date
        if isinstance(posting_date, datetime):
            pd = posting_date.date()
        elif isinstance(posting_date, date):
            pd = posting_date
        else:
            try:
                pd = datetime.strptime(str(posting_date), "%Y-%m-%d").date()
            except Exception:
                pd = date(2026, 4, 30)

        lc = float(lc_amount or 0)
        usd = float(usd_amount or 0)
        jtype = _infer_journal_type(str(acct_type or ""), lc)
        clean_name = str(acct_name).replace("�", "–") if acct_name else None

        db.add(JournalEntry(
            id=str(txn_id),
            entity_id=entity_code,
            counterparty_entity_id=ic_counterparty,
            counterparty_raw=ic_counterparty,
            account_code=str(acct_code),
            account_name=clean_name,
            amount_local=abs(lc),
            currency=CURRENCY_MAP.get((country or "").strip(), "USD"),
            amount_gbp=abs(usd),  # USD as reporting currency
            period=period,
            posting_date=pd,
            description=clean_name,
            journal_type=jtype,
            is_normalised=True,  # entity IDs are direct codes; USD amounts already computed
        ))
        imported += 1

    # Ensure counterparty entities also exist (they may only appear as counterparties)
    all_entries = db.query(JournalEntry).filter(JournalEntry.period == period).all()
    for e in all_entries:
        if e.counterparty_entity_id and not db.get(Entity, e.counterparty_entity_id):
            db.add(Entity(
                id=e.counterparty_entity_id,
                name=e.counterparty_entity_id,
                aliases=[],
                functional_currency="USD",
                ic_agreement_flag=True,
            ))

    # Seed a close calendar entry for the period if missing
    if not db.query(CloseCalendar).filter(CloseCalendar.period == period).first():
        for eid in set(e.entity_id for e in all_entries):
            db.add(CloseCalendar(
                entity_id=eid,
                period=period,
                close_date=date(2026, 5, 8),
                status="open",
            ))

    db.commit()
    return {"period": period, "rows_imported": imported, "rows_skipped": skipped}


# ── TB-only fallback ──────────────────────────────────────────────────────────

def derive_journal_entries_from_tb(db: Session, period: str) -> dict:
    """
    Synthesise JournalEntry records from TrialBalance IC rows when no
    transaction file has been uploaded. Each IC TB line becomes one entry.
    Skips if real (non-synthetic) transaction entries already exist for the period.
    """
    real_count = db.query(JournalEntry).filter(
        JournalEntry.period == period,
        ~JournalEntry.id.like("TB-%"),
    ).count()
    if real_count > 0:
        return {"rows_derived": 0, "period": period, "skipped_reason": "real transaction entries already exist"}

    # Remove any stale synthetic entries first
    db.query(JournalEntry).filter(
        JournalEntry.period == period,
        JournalEntry.id.like("TB-%"),
    ).delete(synchronize_session=False)

    year, month = map(int, period.split("-"))
    posting_date = date(year, month, calendar.monthrange(year, month)[1])

    tb_rows = db.query(TrialBalance).filter(
        TrialBalance.period == period,
        TrialBalance.counterparty_entity_id != None,  # noqa: E711
    ).all()

    derived = 0
    for tb in tb_rows:
        name_lower = (tb.account_name or "").lower()
        if any(w in name_lower for w in ("receivable", "revenue", "income")):
            jtype = "IC_RECEIVABLE"
        elif any(w in name_lower for w in ("payable", "cost", "expense", "recharge")):
            jtype = "IC_PAYABLE"
        else:
            jtype = "IC_RECEIVABLE" if (tb.balance_local or 0) >= 0 else "IC_PAYABLE"

        entry_id = f"TB-{tb.entity_id}-{tb.counterparty_entity_id}-{tb.account_code}-{period}"
        if db.get(JournalEntry, entry_id):
            continue

        db.add(JournalEntry(
            id=entry_id,
            entity_id=tb.entity_id,
            counterparty_entity_id=tb.counterparty_entity_id,
            counterparty_raw=tb.counterparty_entity_id,
            account_code=tb.account_code,
            account_name=tb.account_name,
            amount_local=abs(tb.balance_local or 0),
            currency=tb.currency,
            amount_gbp=abs(tb.balance_gbp) if tb.balance_gbp is not None else None,
            period=period,
            posting_date=posting_date,
            description=f"Derived from TB: {tb.account_name}",
            journal_type=jtype,
            is_normalised=tb.balance_gbp is not None,
        ))
        derived += 1

    db.commit()
    return {"rows_derived": derived, "period": period, "source": "trial_balance"}


# ── FX Rate upload ────────────────────────────────────────────────────────────

HEADER_ALIASES = {
    "from_currency": [
        "from currency", "from", "currency", "currency code", "ccy", "iso",
        "base currency", "base ccy", "source currency", "fx currency",
    ],
    "to_currency": [
        "to currency", "to", "quote currency", "target currency", "reporting currency",
    ],
    "rate": [
        "rate", "fx rate", "exchange rate", "spot rate", "closing rate",
        "period end rate", "rate to gbp", "rate to usd",
    ],
    "effective_date": [
        "date", "effective date", "period end date", "rate date",
        "as of date", "value date", "snapshot date",
    ],
    "period": [
        "period", "accounting period", "fiscal period", "month",
    ],
}


def _normalise_header(value) -> str:
    return str(value or "").lower().strip().replace("_", " ").replace("-", " ")


def _find_col(headers, field):
    for i, h in enumerate(headers):
        if _normalise_header(h) in HEADER_ALIASES[field]:
            return i
    return None


def _parse_date_val(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%d %b %Y", "%b %d %Y"):
            try:
                return datetime.strptime(str(value).strip(), fmt).date()
            except ValueError:
                continue
    return None


def _period_from_date(d: date) -> str:
    return f"{d.year:04d}-{d.month:02d}"


def import_fx_rates(db: Session, file_bytes: bytes, default_period: str = None) -> dict:
    """
    Parse a client-supplied FX rate Excel file and persist to fx_rates.
    Accepts any layout — scans all sheets for a header row with recognisable
    column names, then reads data below it.
    Rates are stored as X -> GBP (or as-is with a warning if to_currency != GBP).
    """
    wb = openpyxl.load_workbook(BytesIO(file_bytes))
    imported = skipped = 0
    warnings = []
    periods_affected = set()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))

        # Find first row that has both a currency col and a rate col
        header_row_idx = None
        col_map = {}
        for i, row in enumerate(all_rows):
            headers = [str(v or "") for v in row]
            fc = _find_col(headers, "from_currency")
            rc = _find_col(headers, "rate")
            if fc is not None and rc is not None:
                header_row_idx = i
                col_map["from_currency"] = fc
                col_map["rate"] = rc
                for field in ("to_currency", "effective_date", "period"):
                    idx = _find_col(headers, field)
                    if idx is not None:
                        col_map[field] = idx
                break

        if header_row_idx is None:
            warnings.append(f"Sheet '{sheet_name}': no recognisable header row — skipped")
            continue

        for row in all_rows[header_row_idx + 1:]:

            def get(field, r=row):
                idx = col_map.get(field)
                return r[idx] if idx is not None and idx < len(r) else None

            from_ccy = str(get("from_currency") or "").strip().upper()
            if not from_ccy or len(from_ccy) > 4:
                skipped += 1
                continue

            try:
                rate = float(get("rate"))
            except (TypeError, ValueError):
                skipped += 1
                continue

            if rate == 0:
                warnings.append(f"Zero rate for {from_ccy} — skipped")
                skipped += 1
                continue

            to_ccy = str(get("to_currency") or "GBP").strip().upper() or "GBP"
            if to_ccy != "GBP":
                warnings.append(
                    f"{from_ccy}/{to_ccy}: to_currency is not GBP — stored as-is; "
                    "normalisation will use this rate but cross-rate triangulation is not applied"
                )

            eff_date = _parse_date_val(get("effective_date")) or date.today()

            raw_period = get("period")
            if raw_period:
                period = str(raw_period).strip()
                m = re.match(r"(\w+)\s+(\d{4})", period)
                if m:
                    months = {
                        "jan": "01", "feb": "02", "mar": "03", "apr": "04",
                        "may": "05", "jun": "06", "jul": "07", "aug": "08",
                        "sep": "09", "oct": "10", "nov": "11", "dec": "12",
                    }
                    period = f"{m.group(2)}-{months.get(m.group(1).lower()[:3], '01')}"
                elif not re.match(r"\d{4}-\d{2}", period):
                    period = _period_from_date(eff_date)
            else:
                period = default_period or _period_from_date(eff_date)

            # Upsert
            db.query(FXRate).filter(
                FXRate.from_currency == from_ccy,
                FXRate.to_currency == to_ccy,
                FXRate.period == period,
            ).delete()

            db.add(FXRate(
                from_currency=from_ccy,
                to_currency=to_ccy,
                rate=rate,
                effective_date=eff_date,
                period=period,
            ))
            periods_affected.add(period)
            imported += 1

    db.commit()
    return {
        "rows_imported": imported,
        "rows_skipped": skipped,
        "periods_affected": sorted(periods_affected),
        "warnings": warnings,
    }
